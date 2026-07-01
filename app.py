import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import openpyxl
import io

st.set_page_config(
    page_title="TCG633 – Pavement Condition Evaluation Tool",
    page_icon="🛣️",
    layout="wide"
)

DEFECT_WEIGHTS = {
    "Alligator (Fatigue) Crack": 2.0,
    "Bleeding/Flushing": 1.0,
    "Block Crack": 1.4,
    "Bumps and Sags": 1.4,
    "Corrugation": 1.5,
    "Depression": 1.6,
    "Edge Crack": 1.1,
    "Lane/Shoulder Drop-off": 1.2,
    "Longitudinal Crack": 1.0,
    "Patching": 1.5,
    "Polished Aggregate": 0.8,
    "Potholes": 2.2,
    "Railroad Crossing": 1.0,
    "Raveling": 1.2,
    "Reflection Crack at Joints": 1.2,
    "Rutting": 1.8,
    "Shoving": 1.8,
    "Spalling of Longitudinal Joint": 1.3,
    "Spalling of Transverse Joints": 1.3,
    "Swells": 1.3,
    "Transverse Crack": 1.1,
    "Weathering": 1.0,
    "Other (Custom)": 1.0,  # fallback weight; user supplies name
}

DEFAULT_CUSTOM_WEIGHT = 1.0  # weight applied to any user-typed custom defect

SEVERITY_FACTORS = {"Low": 0.6, "Medium": 1.0, "High": 1.4}
DEFECT_LIST = list(DEFECT_WEIGHTS.keys())
SEVERITY_LIST = ["Low", "Medium", "High"]

def classify_pci(pci):
    if pci >= 85:
        return "Very Good", "#2ecc71"
    elif pci >= 70:
        return "Good / Satisfactory", "#27ae60"
    elif pci >= 55:
        return "Fair", "#f39c12"
    else:
        return "Poor", "#e74c3c"

def pci_recommendation(pci):
    if pci >= 85:
        return "Routine maintenance (cleaning, grass cutting, minor touch-ups)"
    elif pci >= 70:
        return "Preventive maintenance (crack sealing, local patching)"
    elif pci >= 55:
        return "Surface treatment / Overlay (localized)"
    else:
        return "Major rehabilitation / Reconstruction assessment"

def classify_iri(iri):
    if iri < 2:
        return "Very Good (Smooth)", "#2ecc71"
    elif iri < 3:
        return "Good", "#27ae60"
    elif iri < 4:
        return "Fair", "#f39c12"
    else:
        return "Poor (Rough)", "#e74c3c"

def iri_recommendation(iri):
    if iri < 2:
        return "Routine maintenance"
    elif iri < 3:
        return "Preventive maintenance (localized patching/leveling)"
    elif iri < 4:
        return "Surface treatment / thin overlay"
    else:
        return "Structural overlay / rehabilitation"

CONDITION_RANK = {
    "Very Good": 4, "Very Good (Smooth)": 4,
    "Good / Satisfactory": 3, "Good": 3,
    "Fair": 2,
    "Poor": 1, "Poor (Rough)": 1
}

def hybrid_condition(pci_class, iri_class):
    return pci_class if CONDITION_RANK.get(pci_class, 4) <= CONDITION_RANK.get(iri_class, 4) else iri_class

def hybrid_recommendation(combined):
    rank = CONDITION_RANK.get(combined, 4)
    if rank == 4:
        return "Routine maintenance (cleaning, grass cutting, minor touch-ups)"
    elif rank == 3:
        return "Preventive maintenance (crack sealing, local patching)"
    elif rank == 2:
        return "Surface treatment / Overlay (localized)"
    else:
        return "Major rehabilitation / Reconstruction assessment"

def color_condition(val):
    colors = {
        "Very Good": "background-color: #d5f5e3; color: #1e8449",
        "Good / Satisfactory": "background-color: #d5f5e3; color: #1e8449",
        "Good": "background-color: #d5f5e3; color: #1e8449",
        "Very Good (Smooth)": "background-color: #d5f5e3; color: #1e8449",
        "Fair": "background-color: #fef9e7; color: #d68910",
        "Poor": "background-color: #fadbd8; color: #922b21",
        "Poor (Rough)": "background-color: #fadbd8; color: #922b21"
    }
    return colors.get(val, "")

# ── PHOTO / KEY HELPERS ──
def skey(road_name, section):
    """Composite identity key: (Road Name, Section ID). Lets the same Section ID
    number be reused across different road names."""
    return (str(road_name).strip() if road_name is not None else "-", int(section))

def extract_section_photos(file_obj, sheet_name="PCI_Input", id_col_letter="A"):
    """
    Extract any images embedded/pasted into an Excel sheet and map each one
    to the nearest Section ID above it in the given ID column.
    Excel-uploaded data has no Road Name column, so photos are keyed under "-".
    Returns: {(road_name, section_id): photo_bytes}
    """
    photos = {}
    try:
        file_obj.seek(0)
        wb = openpyxl.load_workbook(file_obj)
        if sheet_name not in wb.sheetnames:
            return photos
        ws = wb[sheet_name]

        # Map worksheet row -> Section ID using the ID column
        row_to_section = {}
        for row in ws.iter_rows():
            for cell in row:
                if cell.column_letter == id_col_letter and isinstance(cell.value, (int, float)):
                    row_to_section[cell.row] = int(cell.value)

        for img in getattr(ws, "_images", []):
            anchor_row = img.anchor._from.row + 1  # openpyxl anchors are 0-indexed
            candidates = [r for r in row_to_section if r <= anchor_row]
            if not candidates:
                continue
            sec = row_to_section[max(candidates)]
            try:
                photos[skey("-", sec)] = img._data()
            except Exception:
                continue
    except Exception:
        pass
    finally:
        try:
            file_obj.seek(0)
        except Exception:
            pass
    return photos

def show_section_photo(key, caption=None, width=260):
    """Display a stored photo for a (road_name, section) key, if one exists."""
    photo = st.session_state.section_photos.get(key)
    if photo:
        label = caption or f"Section {key[1]} — {key[0]} – defect photo"
        st.image(photo, caption=label, width=width)

# ── HEADER ──
st.markdown("""
<div style='background: linear-gradient(135deg, #1a1a2e, #16213e, #0f3460);
            padding: 2rem; border-radius: 12px; margin-bottom: 1.5rem;'>
    <h1 style='color: white; margin:0; font-size: 2rem;'>🛣️ Digital Pavement Condition Evaluation Tool</h1>
    <p style='color: #a0aec0; margin: 0.5rem 0 0 0; font-size: 1rem;'>
        TCG633 – Bridge & Road Maintenance | UiTM Cawangan Sarawak | JKR Standard
    </p>
</div>
""", unsafe_allow_html=True)

# ── SIDEBAR ──
with st.sidebar:
    st.markdown("## ⚙️ Settings")
    mode = st.radio("**Analysis Mode**", ["PCI Only", "IRI Only", "Hybrid (PCI + IRI)"], index=2)
    st.markdown("---")
    st.markdown("### 📋 About")
    st.info("This tool evaluates road pavement condition using PCI (visual defect survey) and IRI (roughness measurement) in accordance with JKR and ASTM D6433 standards.")
    st.markdown("---")

    # ── BENCHMARK REFERENCE TABLE IN SIDEBAR ──
    st.markdown("### 📊 Classification Benchmarks")
    st.markdown("**PCI – Pavement Condition Index**")
    pci_bench = pd.DataFrame({
        "Range": ["85 – 100", "70 – 84", "55 – 69", "0 – 54"],
        "Rating": ["Very Good", "Good / Satisfactory", "Fair", "Poor"],
        "Action": ["Routine maintenance", "Preventive maintenance", "Surface treatment", "Rehabilitation"]
    })
    st.dataframe(pci_bench, use_container_width=True, hide_index=True)

    st.markdown("**IRI – International Roughness Index (m/km)**")
    iri_bench = pd.DataFrame({
        "Range": ["< 2.0", "2.0 – 3.0", "3.0 – 4.0", "> 4.0"],
        "Rating": ["Very Good", "Good", "Fair", "Poor (Rough)"],
        "Action": ["Routine maintenance", "Preventive maintenance", "Surface treatment", "Rehabilitation"]
    })
    st.dataframe(iri_bench, use_container_width=True, hide_index=True)

# ── SESSION STATE ──
if "manual_sections" not in st.session_state:
    st.session_state.manual_sections = []

if "section_photos" not in st.session_state:
    st.session_state.section_photos = {}  # {(road_name, section_id): photo_bytes}

if "edit_target" not in st.session_state:
    st.session_state.edit_target = None  # (road_name, section_id) currently being edited, or None

# ── TABS ──
if mode == "PCI Only":
    tab_labels = ["📖 How to Use", "📥 Data Input", "📊 PCI Analysis"]
elif mode == "IRI Only":
    tab_labels = ["📖 How to Use", "📥 Data Input", "📈 IRI Analysis"]
else:
    tab_labels = ["📖 How to Use", "📥 Data Input", "📊 PCI Analysis", "📈 IRI Analysis", "📊 Dashboard"]

tabs = st.tabs(tab_labels)

# ══════════════════════════════════════════════
# TAB 0 – HOW TO USE
# ══════════════════════════════════════════════
with tabs[0]:
    st.markdown("## 📖 How to Use This App")
    st.markdown("Welcome to the **TCG633 Digital Pavement Condition Evaluation Tool**. Follow the steps below to evaluate your road pavement condition using PCI and/or IRI data.")
    st.markdown("---")

    # STEP 1
    st.markdown("""
<div style='background:#eaf4fb; border-left:5px solid #3498db; padding:1rem 1.2rem; border-radius:8px; margin-bottom:1rem;'>
<h4 style='margin:0 0 0.5rem 0; color:#2471a3;'>Step 1 — Choose Your Analysis Mode</h4>
<p style='margin:0; color:#1a252f;'>
On the <b>left sidebar</b>, select one of three modes:<br><br>
🔵 <b>PCI Only</b> — if you only have visual defect survey data<br>
🔵 <b>IRI Only</b> — if you only have roughness measurement data<br>
🔵 <b>Hybrid (PCI + IRI)</b> — recommended; combines both for a more complete picture
</p>
</div>
""", unsafe_allow_html=True)

    # STEP 2
    st.markdown("""
<div style='background:#eafaf1; border-left:5px solid #2ecc71; padding:1rem 1.2rem; border-radius:8px; margin-bottom:1rem;'>
<h4 style='margin:0 0 0.5rem 0; color:#1e8449;'>Step 2 — Enter Your Data</h4>
<p style='margin:0; color:#1a252f;'>
Go to the <b>📥 Data Input</b> tab. You have two options:<br><br>
📂 <b>Upload Excel File</b> — upload a <code>.xlsx</code> file with sheets named <code>PCI_Input</code> and <code>IRI_Input</code>.<br>
&nbsp;&nbsp;&nbsp;&nbsp;• <code>PCI_Input</code> needs columns: <b>Section ID, Defect Type, Severity, Area Affected (%)</b><br>
&nbsp;&nbsp;&nbsp;&nbsp;• <code>IRI_Input</code> needs columns: <b>Section ID, IRI (m/km)</b><br>
&nbsp;&nbsp;&nbsp;&nbsp;• 📷 If you paste/insert a defect photo into a row of <code>PCI_Input</code>, the app will pick it up and show it against that Section ID.<br><br>
✏️ <b>Enter Data Manually</b> — fill in the form for one section at a time and click <b>Add Section</b>.<br>
&nbsp;&nbsp;&nbsp;&nbsp;• Every field is required: Road Name, Section ID, Defect Type, Severity, Area Affected (%), and IRI — PCI and IRI info must both be filled in before you can save<br>
&nbsp;&nbsp;&nbsp;&nbsp;• The same Section ID number can be reused across different Road Names<br>
&nbsp;&nbsp;&nbsp;&nbsp;• All values are typed directly into the fields (no stepper buttons)<br>
&nbsp;&nbsp;&nbsp;&nbsp;• PCI is calculated immediately after you add or update a section<br>
&nbsp;&nbsp;&nbsp;&nbsp;• You can optionally attach a defect photo (jpg/png) per section<br>
&nbsp;&nbsp;&nbsp;&nbsp;• Pick any section from the list below the table to <b>Edit</b> or <b>Delete</b> it, or use <b>Clear All</b> to start over
</p>
</div>
""", unsafe_allow_html=True)

    # STEP 3
    st.markdown("""
<div style='background:#fef9e7; border-left:5px solid #f39c12; padding:1rem 1.2rem; border-radius:8px; margin-bottom:1rem;'>
<h4 style='margin:0 0 0.5rem 0; color:#d68910;'>Step 3 — View the Analysis</h4>
<p style='margin:0; color:#1a252f;'>
Once data is entered, navigate to the analysis tabs:<br><br>
📊 <b>PCI Analysis</b> — shows PCI scores per section, a condition distribution pie chart, and a full computation table<br>
📈 <b>IRI Analysis</b> — shows the IRI roughness profile, condition zones, and results table<br>
📊 <b>Dashboard</b> — combines both into one overall condition rating per section, with a section filter and a maintenance priority plan
</p>
</div>
""", unsafe_allow_html=True)

    # STEP 4
    st.markdown("""
<div style='background:#fdf2f8; border-left:5px solid #8e44ad; padding:1rem 1.2rem; border-radius:8px; margin-bottom:1rem;'>
<h4 style='margin:0 0 0.5rem 0; color:#6c3483;'>Step 4 — Read the Results</h4>
<p style='margin:0; color:#1a252f;'>
Each section is rated using colour-coded conditions:<br><br>
🟢 <b>Very Good (PCI 85–100 / IRI &lt; 2.0)</b> — Routine maintenance only<br>
🟢 <b>Good / Satisfactory (PCI 70–84 / IRI 2.0–3.0)</b> — Preventive maintenance (crack sealing, patching)<br>
🟡 <b>Fair (PCI 55–69 / IRI 3.0–4.0)</b> — Surface treatment or localized overlay needed<br>
🔴 <b>Poor (PCI &lt; 55 / IRI &gt; 4.0)</b> — Major rehabilitation or reconstruction required<br><br>
In <b>Hybrid mode</b>, the combined condition is always the <b>worse</b> of PCI and IRI — the conservative approach used in JKR assessments.
</p>
</div>
""", unsafe_allow_html=True)

    # STEP 5
    st.markdown("""
<div style='background:#f2f3f4; border-left:5px solid #7f8c8d; padding:1rem 1.2rem; border-radius:8px; margin-bottom:1rem;'>
<h4 style='margin:0 0 0.5rem 0; color:#515a5a;'>Step 5 — Filter & Export</h4>
<p style='margin:0; color:#1a252f;'>
In the <b>📊 Dashboard</b> tab, use the section filter to choose which section(s) you want to view — pick one, several, or all of them.
Then click <b>⬇️ Download Summary as CSV</b> to save your (filtered) results for reporting or further analysis.
</p>
</div>
""", unsafe_allow_html=True)

    st.markdown("---")
    st.markdown("### ❓ Frequently Asked Questions")

    with st.expander("What is PCI?"):
        st.markdown("""
**Pavement Condition Index (PCI)** is a numerical rating from **0 to 100** that describes the surface condition of a road section based on visual inspection of defects.
- It is defined by **ASTM D6433** and used worldwide including under **JKR Malaysia** standards.
- Inspectors identify defect types (e.g. cracks, potholes, raveling), their severity (Low/Medium/High), and the area affected.
- The app calculates: `PCI = 100 − (Area % × Defect Weight × Severity Factor)`
""")

    with st.expander("What is IRI?"):
        st.markdown("""
**International Roughness Index (IRI)** measures road surface roughness in **m/km** (metres per kilometre).
- A lower IRI means a smoother road. A higher IRI means a rougher, bumpier surface.
- It is typically measured using a profilometer or smartphone-based tools.
- IRI < 2.0 is considered very smooth (like a new highway); IRI > 4.0 indicates a road needing urgent attention.
""")

    with st.expander("What does Hybrid / Dashboard mode do?"):
        st.markdown("""
The Dashboard combines PCI and IRI into a single **Combined Condition** rating.
- The app compares the condition class of PCI and IRI for each section.
- It always picks the **worse** of the two — this is the conservative approach recommended by JKR.
- Example: If PCI = Very Good but IRI = Fair, the Combined Condition = **Fair**.
- You can filter the dashboard to focus on just the section(s) you're interested in.
""")

    with st.expander("Can I enter my own defect types?"):
        st.markdown("""
Yes! In the manual entry form, the **Defect Type** dropdown includes 22 standard ASTM D6433 defect types.
If your defect is not listed, select **Other (Custom)** and type the name manually.
A default weight of **1.0** will be applied — the PCI result will be an estimate.
""")

    with st.expander("How do defect photos work?"):
        st.markdown("""
There are two ways to attach a defect photo to a section:
- **Excel upload**: paste or insert a picture directly onto a row of the `PCI_Input` sheet, near that section's data. The app finds the nearest Section ID above the picture and links it automatically.
- **Manual entry**: use the photo uploader in the entry form to attach a jpg/png for that section.

Photos then appear in the Data Input table, the PCI Analysis table, and the Dashboard's Maintenance Priority Plan.
""")

    with st.expander("Why does my Excel file not load all sections?"):
        st.markdown("""
Make sure your Excel file:
- Has sheets named exactly **`PCI_Input`** and **`IRI_Input`** (case-sensitive)
- Has a header row with columns: **Section ID**, **Defect Type**, **Severity**, **Area Affected (%)**
- Uses **Section ID** as a plain number (e.g. 1, 2, 3 — not "Section 1")
- Uses **Severity** values of exactly: `Low`, `Medium`, or `High`

If a row is skipped, the app will show a warning message explaining why.
""")


# ══════════════════════════════════════════════
# TAB 1 – DATA INPUT
# ══════════════════════════════════════════════
with tabs[1]:
    st.markdown("## 📥 Data Input")

    input_method = st.radio(
        "**Choose Input Method**",
        ["📂 Upload Excel File", "✏️ Enter Data Manually"],
        horizontal=True
    )
    st.markdown("---")

    pci_data = []
    iri_data = []

    # ══════════════════════════════════════════
    # OPTION A: UPLOAD EXCEL
    # ══════════════════════════════════════════
    if input_method == "📂 Upload Excel File":
        st.markdown("### 📂 Upload Your Excel Dataset")
        st.caption("Upload your TCG633 Excel file. The app will read the PCI_Input and IRI_Input sheets automatically, and pick up any defect photos pasted into PCI_Input.")

        if "excel_uploader_key" not in st.session_state:
            st.session_state.excel_uploader_key = 0
        if "last_excel_signature" not in st.session_state:
            st.session_state.last_excel_signature = None

        col_up, col_reset = st.columns([4, 1])
        with col_up:
            uploaded_file = st.file_uploader(
                "Drop your Excel file here or click to browse",
                type=["xlsx", "xls"],
                help="Must contain sheets named 'PCI_Input' and 'IRI_Input'",
                key=f"excel_file_{st.session_state.excel_uploader_key}"
            )
        with col_reset:
            st.markdown("<div style='height:1.9rem'></div>", unsafe_allow_html=True)
            if st.button("🔄 Reset Upload", use_container_width=True,
                        help="Click this before re-uploading an edited file, especially if it has the SAME filename as before — browsers sometimes won't register a re-upload otherwise."):
                st.session_state.excel_uploader_key += 1
                st.session_state.last_excel_signature = None
                # Clear photos that came from a previous Excel upload (keyed under Road Name "-")
                st.session_state.section_photos = {k: v for k, v in st.session_state.section_photos.items() if k[0] != "-"}
                st.rerun()

        if uploaded_file is not None:
            try:
                # Detect whether this is genuinely new file content vs. the same file re-rendering
                file_signature = (uploaded_file.name, uploaded_file.size, hash(uploaded_file.getvalue()))
                is_new_file = file_signature != st.session_state.last_excel_signature
                if is_new_file:
                    st.session_state.last_excel_signature = file_signature
                    # Wipe out any photos from a PREVIOUS Excel upload so stale ones don't linger
                    st.session_state.section_photos = {k: v for k, v in st.session_state.section_photos.items() if k[0] != "-"}

                st.caption(f"📄 Currently loaded: **{uploaded_file.name}** ({uploaded_file.size:,} bytes). "
                          f"If you edited this file and it still shows old values below, click **🔄 Reset Upload** above and upload it again.")

                xl = pd.ExcelFile(uploaded_file)
                available_sheets = xl.sheet_names
                st.success(f"✅ File uploaded! Sheets found: {', '.join(available_sheets)}")

                # ── Extract embedded defect photos from PCI_Input ──
                photo_bytes = io.BytesIO(uploaded_file.getvalue())
                extracted_photos = extract_section_photos(photo_bytes, sheet_name="PCI_Input")
                if extracted_photos:
                    st.session_state.section_photos.update(extracted_photos)
                    st.success(f"📷 Found {len(extracted_photos)} defect photo(s) in PCI_Input, linked to sections: {', '.join(str(k[1]) for k in sorted(extracted_photos))}")

                col_pci, col_iri = st.columns(2)

                with col_pci:
                    if "PCI_Input" in available_sheets:
                        df_pci_input = xl.parse("PCI_Input", header=None)
                        header_row = None
                        for idx, row in df_pci_input.iterrows():
                            if any(str(cell).strip() == "Section ID" for cell in row):
                                header_row = idx
                                break
                        if header_row is not None:
                            df_pci_input.columns = df_pci_input.iloc[header_row]
                            df_pci_input = df_pci_input.iloc[header_row+1:].reset_index(drop=True)
                            df_pci_input = df_pci_input.dropna(subset=["Section ID"])
                            df_pci_input = df_pci_input[df_pci_input["Section ID"].astype(str).str.strip().str.match(r'^\d+(\.\d+)?$')]
                            st.markdown("### 🔍 PCI Data from Excel")
                            st.dataframe(df_pci_input[["Section ID","Defect Type","Severity","Area Affected (%)"]].reset_index(drop=True),
                                         use_container_width=True, hide_index=True)
                            skipped_pci = []
                            for _, row in df_pci_input.iterrows():
                                try:
                                    section = int(float(str(row["Section ID"]).strip()))
                                    defect = str(row["Defect Type"]).strip()
                                    severity = str(row["Severity"]).strip().capitalize()
                                    area = float(str(row["Area Affected (%)"]).strip()) if str(row["Area Affected (%)"]).strip() not in ["nan",""] else 0.0
                                    if severity not in SEVERITY_FACTORS:
                                        skipped_pci.append(f"S{section}: unrecognised severity '{severity}' (use Low/Medium/High)")
                                        continue
                                    pci_data.append({"Section": section, "Defect Type": defect,
                                                     "Severity": severity, "Area (%)": area})
                                except Exception as ex:
                                    skipped_pci.append(f"Row skipped: {ex}")
                                    continue
                            if skipped_pci:
                                st.warning("\u26a0\ufe0f Some PCI rows were skipped:\n" + "\n".join(f"- {s}" for s in skipped_pci))

                            # ── Photo gallery for this upload ──
                            if extracted_photos:
                                st.markdown("#### 📷 Defect Photos")
                                gallery_cols = st.columns(min(4, len(extracted_photos)))
                                for i, (key, img) in enumerate(sorted(extracted_photos.items())):
                                    with gallery_cols[i % len(gallery_cols)]:
                                        st.image(img, caption=f"Section {key[1]}", use_container_width=True)
                        else:
                            st.warning("Could not find header row in PCI_Input sheet.")
                    else:
                        st.warning("PCI_Input sheet not found in uploaded file.")

                with col_iri:
                    if "IRI_Input" in available_sheets:
                        df_iri_input = xl.parse("IRI_Input", header=None)
                        header_row_iri = None
                        for idx, row in df_iri_input.iterrows():
                            if any(str(cell).strip() == "Section ID" for cell in row):
                                header_row_iri = idx
                                break
                        if header_row_iri is not None:
                            df_iri_input.columns = df_iri_input.iloc[header_row_iri]
                            df_iri_input = df_iri_input.iloc[header_row_iri+1:].reset_index(drop=True)
                            df_iri_input = df_iri_input.dropna(subset=["Section ID"])
                            df_iri_input = df_iri_input[df_iri_input["Section ID"].astype(str).str.strip().str.match(r'^\d+(\.\d+)?$')]
                            df_iri_input["Section ID"] = df_iri_input["Section ID"].astype(float).astype(int)
                            df_iri_input["IRI (m/km)"] = pd.to_numeric(df_iri_input["IRI (m/km)"], errors="coerce")
                            iri_avg = df_iri_input.groupby("Section ID")["IRI (m/km)"].mean().reset_index()
                            iri_avg.columns = ["Section", "IRI (m/km)"]
                            st.markdown("### 📏 IRI Data from Excel")
                            st.dataframe(iri_avg, use_container_width=True, hide_index=True)
                            for _, row in iri_avg.iterrows():
                                iri_data.append({"Section": int(row["Section"]),
                                                 "IRI (m/km)": float(row["IRI (m/km)"])})
                        else:
                            st.warning("Could not find header row in IRI_Input sheet.")
                    else:
                        st.warning("IRI_Input sheet not found in uploaded file.")

                if pci_data and iri_data:
                    st.success("✅ Data loaded successfully! Navigate to the analysis tabs to view results.")
                else:
                    st.error("❌ Could not load data. Please check your Excel file format.")

            except Exception as e:
                st.error(f"❌ Error reading file: {str(e)}")
        else:
            st.info("👆 Please upload your Excel file to proceed.")

    # ══════════════════════════════════════════
    # OPTION B: MANUAL INPUT — FORM TEMPLATE
    # ══════════════════════════════════════════
    else:
        st.markdown("### ✏️ Add / Edit a Section")

        edit_data = None
        if st.session_state.edit_target is not None:
            target_road, target_sec = st.session_state.edit_target
            edit_data = next((s for s in st.session_state.manual_sections
                               if s["Road Name"] == target_road and s["Section"] == target_sec), None)

        if edit_data:
            st.info(f"✏️ Editing **Section {edit_data['Section']} — {edit_data['Road Name']}**. Update the fields below and click **Update Section**.")
            if st.button("✖️ Cancel Edit"):
                st.session_state.edit_target = None
                st.rerun()
        else:
            st.caption("Fill in every field, then click **Add Section**. Road Name, Section ID, PCI info, and IRI info are all required before you can proceed. "
                       "The same Section ID can be reused as long as the Road Name is different.")

        default_defect_index = 0
        if edit_data:
            default_defect_index = DEFECT_LIST.index("Other (Custom)") if edit_data.get("Is Custom") \
                else (DEFECT_LIST.index(edit_data["Defect Type"]) if edit_data["Defect Type"] in DEFECT_LIST else 0)
        default_severity_index = SEVERITY_LIST.index(edit_data["Severity"]) if edit_data else 0

        # ── ENTRY FORM ──
        with st.form("section_form", clear_on_submit=not edit_data):
            st.markdown("#### 📋 Section Entry Form")

            st.markdown("**🛣️ Location Info**")
            loc1, loc2 = st.columns(2)
            with loc1:
                f_road_name = st.text_input("Road Name", value=edit_data["Road Name"] if edit_data else "",
                                            placeholder="e.g. Jalan Datuk Mohd Musa")
            with loc2:
                f_section_str = st.text_input("Section ID (number)", value=str(edit_data["Section"]) if edit_data else "",
                                              placeholder="e.g. 1")
                st.caption("ℹ️ The same Section ID can be reused for a different Road Name.")

            col1, col2 = st.columns(2)

            with col1:
                st.markdown("**🔍 PCI – Defect Info**")
                f_defect = st.selectbox("Defect Type", DEFECT_LIST, index=default_defect_index,
                                        help="Select the primary defect observed in this section. Choose 'Other (Custom)' if your defect is not listed.")
                f_custom_defect = ""
                if f_defect == "Other (Custom)":
                    f_custom_defect = st.text_input(
                        "Custom Defect Name",
                        value=edit_data["Defect Type"] if edit_data and edit_data.get("Is Custom") else "",
                        placeholder="e.g. Depression, Delamination, Joint Failure…",
                        help="Type in the exact defect name. A default weight of 1.0 will be applied."
                    )
                    st.caption("ℹ️ Custom defects use a default weighting of **1.0**. The PCI result is an estimate.")
                f_severity = st.selectbox("Severity Level", SEVERITY_LIST, index=default_severity_index,
                                          help="Low = minor, Medium = moderate, High = severe")
                f_area_str = st.text_input("Area Affected (%)", value=str(edit_data["Area (%)"]) if edit_data else "",
                                           placeholder="e.g. 5 (required for PCI)")
                f_photo = st.file_uploader("📷 Defect Photo (optional)", type=["jpg", "jpeg", "png"],
                                           help="Attach a site photo of the observed defect for this section")

            with col2:
                st.markdown("**📏 IRI – Roughness Info**")
                f_iri_str = st.text_input("IRI Value (m/km)", value=str(edit_data["IRI (m/km)"]) if edit_data else "",
                                          placeholder="e.g. 2.0 (required for IRI)")
                if f_photo is not None:
                    st.image(f_photo, caption="Photo preview", use_container_width=True)
                elif edit_data and skey(edit_data["Road Name"], edit_data["Section"]) in st.session_state.section_photos:
                    st.image(st.session_state.section_photos[skey(edit_data["Road Name"], edit_data["Section"])],
                             caption="Current photo (upload a new one to replace)", use_container_width=True)

                # Quick reference hint inside form
                st.markdown("""
<div style='background:#f0f4ff; border-left: 4px solid #3498db;
            padding: 0.8rem 1rem; border-radius: 6px; font-size: 0.85rem; margin-top: 0.5rem;'>
<b>Quick Reference</b><br>
🟢 IRI &lt; 2.0 → Very Good<br>
🟡 IRI 2.0–3.0 → Good<br>
🟠 IRI 3.0–4.0 → Fair<br>
🔴 IRI &gt; 4.0 → Poor
</div>
""", unsafe_allow_html=True)

            st.markdown("---")
            submit_label = "💾 Update Section" if edit_data else "➕ Add Section"
            submitted = st.form_submit_button(submit_label, use_container_width=True, type="primary")

        if submitted:
            # ── Validate that every field is filled in ──
            missing = []
            if not f_road_name.strip():
                missing.append("Road Name")
            if not f_section_str.strip():
                missing.append("Section ID")

            section_val = None
            if f_section_str.strip():
                try:
                    section_val = int(float(f_section_str.strip()))
                except ValueError:
                    missing.append("Section ID (must be a number)")

            if f_defect == "Other (Custom)" and not f_custom_defect.strip():
                missing.append("Custom Defect Name")

            if not f_area_str.strip():
                missing.append("Area Affected (%) — required for PCI")
            area_val = None
            if f_area_str.strip():
                try:
                    area_val = float(f_area_str.strip())
                    if not (0 <= area_val <= 100):
                        missing.append("Area Affected (%) must be between 0 and 100")
                except ValueError:
                    missing.append("Area Affected (%) must be a number")

            if not f_iri_str.strip():
                missing.append("IRI Value (m/km) — required for IRI")
            iri_val = None
            if f_iri_str.strip():
                try:
                    iri_val = float(f_iri_str.strip())
                    if iri_val < 0:
                        missing.append("IRI Value (m/km) must be 0 or higher")
                except ValueError:
                    missing.append("IRI Value (m/km) must be a number")

            if missing:
                st.warning("⚠️ Please fix the following before saving — PCI and IRI info are both required:\n" +
                           "\n".join(f"- {m}" for m in missing))
            else:
                actual_defect = f_custom_defect.strip() if f_defect == "Other (Custom)" else f_defect
                new_road = f_road_name.strip()
                new_key = skey(new_road, section_val)

                # Duplicate check: same Road Name + Section ID combo (ignore the entry being edited).
                # A Section ID CAN be reused as long as the Road Name is different.
                existing_keys = [skey(s["Road Name"], s["Section"]) for s in st.session_state.manual_sections
                                 if not edit_data or skey(s["Road Name"], s["Section"]) != skey(edit_data["Road Name"], edit_data["Section"])]
                if new_key in existing_keys:
                    st.warning(f"⚠️ Section {section_val} on **{new_road}** already exists. Please use a different Section ID, a different Road Name, or edit the existing entry instead.")
                else:
                    new_entry = {
                        "Section": section_val,
                        "Road Name": new_road,
                        "Defect Type": actual_defect,
                        "Is Custom": f_defect == "Other (Custom)",
                        "Severity": f_severity,
                        "Area (%)": area_val,
                        "IRI (m/km)": iri_val,
                    }

                    if edit_data:
                        old_key = skey(edit_data["Road Name"], edit_data["Section"])
                        idx = next(i for i, s in enumerate(st.session_state.manual_sections)
                                  if skey(s["Road Name"], s["Section"]) == old_key)
                        st.session_state.manual_sections[idx] = new_entry
                        if old_key != new_key and old_key in st.session_state.section_photos:
                            st.session_state.section_photos[new_key] = st.session_state.section_photos.pop(old_key)
                        st.session_state.edit_target = None
                    else:
                        st.session_state.manual_sections.append(new_entry)

                    if f_photo is not None:
                        st.session_state.section_photos[new_key] = f_photo.getvalue()

                    # ── Calculate PCI immediately ──
                    w = DEFECT_WEIGHTS.get(actual_defect, DEFAULT_CUSTOM_WEIGHT)
                    sf = SEVERITY_FACTORS.get(f_severity, 1.0)
                    pci_now = max(0, 100 - area_val * w * sf)
                    pci_cond, _ = classify_pci(pci_now)
                    iri_cond, _ = classify_iri(iri_val)
                    action = "updated" if edit_data else "added"
                    st.success(f"✅ Section {section_val} {action}! PCI = **{pci_now:.1f}** ({pci_cond}) | IRI = **{iri_val:.2f}** ({iri_cond})")
                    st.rerun()

        # ── SECTIONS TABLE ──
        st.markdown("---")
        st.markdown("### 📊 Sections Entered So Far")

        if not st.session_state.manual_sections:
            st.info("No sections added yet. Fill in the form above and click **Add Section** to get started.")
        else:
            df_manual = pd.DataFrame(st.session_state.manual_sections).sort_values("Section").reset_index(drop=True)

            def quick_pci(row):
                w = DEFECT_WEIGHTS.get(row["Defect Type"], DEFAULT_CUSTOM_WEIGHT)
                sf = SEVERITY_FACTORS.get(row["Severity"], 1.0)
                return round(max(0, 100 - row["Area (%)"] * w * sf), 1)

            df_manual["PCI"] = df_manual.apply(quick_pci, axis=1)
            df_manual["PCI Rating"] = df_manual["PCI"].apply(lambda x: classify_pci(x)[0])
            df_manual["IRI Rating"] = df_manual["IRI (m/km)"].apply(lambda x: classify_iri(x)[0])
            df_manual["Photo"] = df_manual.apply(
                lambda row: "📷 Yes" if skey(row["Road Name"], row["Section"]) in st.session_state.section_photos else "—", axis=1)

            display_cols = ["Section", "Road Name", "Defect Type", "Severity", "Area (%)",
                            "IRI (m/km)", "PCI", "PCI Rating", "IRI Rating", "Photo"]

            styled_manual = df_manual[display_cols].style.map(
                color_condition, subset=["PCI Rating", "IRI Rating"]
            )
            st.dataframe(styled_manual, use_container_width=True, hide_index=True)

            st.caption(f"**{len(df_manual)} section(s) entered.** PCI is calculated automatically as soon as a section is added or updated.")

            sections_with_photo = [skey(s["Road Name"], s["Section"]) for s in st.session_state.manual_sections
                                   if skey(s["Road Name"], s["Section"]) in st.session_state.section_photos]
            if sections_with_photo:
                with st.expander(f"📷 View Defect Photos ({len(sections_with_photo)})"):
                    gallery_cols = st.columns(min(4, len(sections_with_photo)))
                    for i, key in enumerate(sections_with_photo):
                        with gallery_cols[i % len(gallery_cols)]:
                            st.image(st.session_state.section_photos[key], caption=f"Section {key[1]} — {key[0]}", use_container_width=True)

            # ── Edit / Delete by picking from a list ──
            st.markdown("#### ✏️ Edit or 🗑️ Delete a Section")
            section_options = {
                skey(s["Road Name"], s["Section"]): f"S{s['Section']} — {s['Road Name'] or '-'}"
                for s in st.session_state.manual_sections
            }
            picked = st.selectbox(
                "Choose a section",
                options=list(section_options.keys()),
                format_func=lambda x: section_options[x]
            )
            col_edit, col_delete = st.columns(2)
            with col_edit:
                if st.button("✏️ Edit Selected Section", use_container_width=True):
                    st.session_state.edit_target = picked
                    st.rerun()
            with col_delete:
                if st.button("🗑️ Delete Selected Section", use_container_width=True):
                    st.session_state.manual_sections = [
                        s for s in st.session_state.manual_sections if skey(s["Road Name"], s["Section"]) != picked
                    ]
                    st.session_state.section_photos.pop(picked, None)
                    if st.session_state.edit_target == picked:
                        st.session_state.edit_target = None
                    st.success(f"Deleted Section {picked[1]} — {picked[0]}.")
                    st.rerun()

            if st.button("🗑️ Clear All Sections", type="secondary"):
                st.session_state.manual_sections = []
                st.session_state.section_photos = {}
                st.session_state.edit_target = None
                st.rerun()

        # Build pci_data / iri_data from session state for analysis tabs
        for s in st.session_state.manual_sections:
            pci_data.append({
                "Section": s["Section"],
                "Road Name": s.get("Road Name", "-"),
                "Defect Type": s["Defect Type"],
                "Severity": s["Severity"],
                "Area (%)": s["Area (%)"],
            })
            iri_data.append({
                "Section": s["Section"],
                "Road Name": s.get("Road Name", "-"),
                "IRI (m/km)": s["IRI (m/km)"],
            })

        if st.session_state.manual_sections:
            st.success("✅ Navigate to the Analysis tabs to view full results.")

# ── COMPUTE ──
pci_results = []
for row in pci_data:
    w = DEFECT_WEIGHTS.get(row["Defect Type"], DEFAULT_CUSTOM_WEIGHT)
    sf = SEVERITY_FACTORS.get(row["Severity"], 1.0)
    deduct = row["Area (%)"] * w * sf
    pci = max(0, 100 - deduct)
    condition, color = classify_pci(pci)
    pci_results.append({
        "Section": row["Section"], "Road Name": row.get("Road Name", "-"),
        "Defect Type": row["Defect Type"],
        "Severity": row["Severity"], "Area (%)": row["Area (%)"],
        "Weighting": w, "Severity Factor": sf, "Deduct Value": round(deduct, 2),
        "PCI": round(pci, 2), "Condition": condition, "Color": color,
        "Recommendation": pci_recommendation(pci)
    })

iri_results = []
for row in iri_data:
    condition, color = classify_iri(row["IRI (m/km)"])
    iri_results.append({
        "Section": row["Section"], "Road Name": row.get("Road Name", "-"),
        "IRI (m/km)": row["IRI (m/km)"],
        "Condition": condition, "Color": color,
        "Recommendation": iri_recommendation(row["IRI (m/km)"])
    })

df_pci = pd.DataFrame(pci_results) if pci_results else pd.DataFrame()
df_iri = pd.DataFrame(iri_results) if iri_results else pd.DataFrame()


def benchmark_table_pci():
    """Render a color-coded PCI benchmark reference card."""
    st.markdown("""
<div style='margin-bottom:1rem;'>
<b>📖 PCI Benchmark Reference</b>
<table style='width:100%; border-collapse:collapse; font-size:0.88rem; margin-top:0.5rem;'>
<tr style='background:#2c3e50; color:white;'>
  <th style='padding:6px 10px; text-align:left;'>PCI Range</th>
  <th style='padding:6px 10px; text-align:left;'>Rating</th>
  <th style='padding:6px 10px; text-align:left;'>What it means</th>
  <th style='padding:6px 10px; text-align:left;'>Recommended Action</th>
</tr>
<tr style='background:#d5f5e3; color:#1e8449;'>
  <td style='padding:6px 10px;'>85 – 100</td><td style='padding:6px 10px;'>🟢 Very Good</td>
  <td style='padding:6px 10px;'>Pavement in excellent condition, minor or no defects</td>
  <td style='padding:6px 10px;'>Routine maintenance (cleaning, grass cutting)</td>
</tr>
<tr style='background:#eafaf1; color:#1e8449;'>
  <td style='padding:6px 10px;'>70 – 84</td><td style='padding:6px 10px;'>🟢 Good / Satisfactory</td>
  <td style='padding:6px 10px;'>Slight defects, still functional with good ride quality</td>
  <td style='padding:6px 10px;'>Preventive maintenance (crack sealing, local patching)</td>
</tr>
<tr style='background:#fef9e7; color:#d68910;'>
  <td style='padding:6px 10px;'>55 – 69</td><td style='padding:6px 10px;'>🟡 Fair</td>
  <td style='padding:6px 10px;'>Visible defects affecting ride comfort, early deterioration</td>
  <td style='padding:6px 10px;'>Surface treatment / localized overlay</td>
</tr>
<tr style='background:#fadbd8; color:#922b21;'>
  <td style='padding:6px 10px;'>0 – 54</td><td style='padding:6px 10px;'>🔴 Poor</td>
  <td style='padding:6px 10px;'>Severe defects, significant structural or surface failure</td>
  <td style='padding:6px 10px;'>Major rehabilitation / reconstruction assessment</td>
</tr>
</table>
</div>
""", unsafe_allow_html=True)


def benchmark_table_iri():
    """Render a color-coded IRI benchmark reference card."""
    st.markdown("""
<div style='margin-bottom:1rem;'>
<b>📖 IRI Benchmark Reference</b>
<table style='width:100%; border-collapse:collapse; font-size:0.88rem; margin-top:0.5rem;'>
<tr style='background:#2c3e50; color:white;'>
  <th style='padding:6px 10px; text-align:left;'>IRI (m/km)</th>
  <th style='padding:6px 10px; text-align:left;'>Rating</th>
  <th style='padding:6px 10px; text-align:left;'>What it means</th>
  <th style='padding:6px 10px; text-align:left;'>Recommended Action</th>
</tr>
<tr style='background:#d5f5e3; color:#1e8449;'>
  <td style='padding:6px 10px;'>&lt; 2.0</td><td style='padding:6px 10px;'>🟢 Very Good</td>
  <td style='padding:6px 10px;'>Very smooth ride, like a new highway surface</td>
  <td style='padding:6px 10px;'>Routine maintenance</td>
</tr>
<tr style='background:#eafaf1; color:#1e8449;'>
  <td style='padding:6px 10px;'>2.0 – 3.0</td><td style='padding:6px 10px;'>🟢 Good</td>
  <td style='padding:6px 10px;'>Slightly rough but comfortable; minor surface variation</td>
  <td style='padding:6px 10px;'>Preventive maintenance (localized patching/leveling)</td>
</tr>
<tr style='background:#fef9e7; color:#d68910;'>
  <td style='padding:6px 10px;'>3.0 – 4.0</td><td style='padding:6px 10px;'>🟡 Fair</td>
  <td style='padding:6px 10px;'>Noticeable roughness, causes driver/passenger discomfort</td>
  <td style='padding:6px 10px;'>Surface treatment / thin overlay</td>
</tr>
<tr style='background:#fadbd8; color:#922b21;'>
  <td style='padding:6px 10px;'>&gt; 4.0</td><td style='padding:6px 10px;'>🔴 Poor (Rough)</td>
  <td style='padding:6px 10px;'>Severely rough, vehicle damage risk, urgent attention needed</td>
  <td style='padding:6px 10px;'>Structural overlay / rehabilitation</td>
</tr>
</table>
</div>
""", unsafe_allow_html=True)


# ══════════════════════════════════════════════
# PCI TAB
# ══════════════════════════════════════════════
if mode in ["PCI Only", "Hybrid (PCI + IRI)"]:
    pci_tab = tabs[2]
    with pci_tab:
        st.markdown("## 📊 PCI Analysis Results")
        if df_pci.empty:
            st.warning("⚠️ No PCI data available. Please upload a file or enter data manually in the Data Input tab.")
        else:
            avg_pci = df_pci["PCI"].mean()
            worst = df_pci.loc[df_pci["PCI"].idxmin()]
            best = df_pci.loc[df_pci["PCI"].idxmax()]
            n = len(df_pci)
            k1, k2, k3, k4 = st.columns(4)
            k1.metric("Average PCI", f"{avg_pci:.1f}")
            k2.metric("Best Section", f"S{int(best['Section'])} ({best['PCI']:.1f})")
            k3.metric("Worst Section", f"S{int(worst['Section'])} ({worst['PCI']:.1f})")
            k4.metric("Sections Needing Attention", f"{len(df_pci[df_pci['Condition'].isin(['Poor','Fair'])])}/{n}")

            st.markdown("---")

            benchmark_table_pci()

            st.markdown("---")
            col1, col2 = st.columns([3, 2])
            with col1:
                st.markdown("### PCI Score by Section")
                fig_bar = go.Figure()
                for _, row in df_pci.iterrows():
                    fig_bar.add_trace(go.Bar(x=[f"S{int(row['Section'])}"], y=[row["PCI"]],
                        marker_color=row["Color"], showlegend=False,
                        text=[f"{row['PCI']:.1f}"], textposition="outside"))

                fig_bar.add_hrect(y0=85, y1=110, fillcolor="#2ecc71", opacity=0.07, line_width=0)
                fig_bar.add_hrect(y0=70, y1=85, fillcolor="#27ae60", opacity=0.07, line_width=0)
                fig_bar.add_hrect(y0=55, y1=70, fillcolor="#f39c12", opacity=0.08, line_width=0)
                fig_bar.add_hrect(y0=0, y1=55, fillcolor="#e74c3c", opacity=0.07, line_width=0)

                fig_bar.add_hline(y=85, line_dash="dash", line_color="#2ecc71",
                                  annotation_text="Very Good (85)", annotation_position="right")
                fig_bar.add_hline(y=70, line_dash="dash", line_color="#27ae60",
                                  annotation_text="Good (70)", annotation_position="right")
                fig_bar.add_hline(y=55, line_dash="dash", line_color="#f39c12",
                                  annotation_text="Fair (55)", annotation_position="right")

                fig_bar.update_layout(yaxis_range=[0, 115], yaxis_title="PCI Score",
                                      xaxis_title="Section", height=420,
                                      plot_bgcolor="rgba(0,0,0,0)", paper_bgcolor="rgba(0,0,0,0)")
                st.plotly_chart(fig_bar, use_container_width=True)

            with col2:
                st.markdown("### Condition Distribution")
                cond_counts = df_pci["Condition"].value_counts().reset_index()
                cond_counts.columns = ["Condition", "Count"]
                fig_pie = px.pie(cond_counts, names="Condition", values="Count", color="Condition",
                                 color_discrete_map={"Very Good": "#2ecc71", "Good / Satisfactory": "#27ae60",
                                                     "Fair": "#f39c12", "Poor": "#e74c3c"}, hole=0.4)
                fig_pie.update_layout(height=420, paper_bgcolor="rgba(0,0,0,0)")
                st.plotly_chart(fig_pie, use_container_width=True)

            st.markdown("### 📋 PCI Computation Table")
            display_pci = df_pci[["Section", "Road Name", "Defect Type", "Severity", "Area (%)", "Weighting",
                                   "Severity Factor", "Deduct Value", "PCI", "Condition", "Recommendation"]].copy()
            styled = display_pci.style.map(color_condition, subset=["Condition"])
            st.dataframe(styled, use_container_width=True, hide_index=True)

            sections_with_photos = [skey(row["Road Name"], row["Section"]) for _, row in df_pci.iterrows()
                                    if skey(row["Road Name"], row["Section"]) in st.session_state.section_photos]
            if sections_with_photos:
                st.markdown("### 📷 Defect Photos by Section")
                for key in sections_with_photos:
                    row = df_pci[(df_pci["Road Name"] == key[0]) & (df_pci["Section"] == key[1])].iloc[0]
                    with st.expander(f"Section {key[1]} — {key[0]} — {row['Condition']} (PCI {row['PCI']:.1f})"):
                        c1, c2 = st.columns([1, 2])
                        with c1:
                            show_section_photo(key)
                        with c2:
                            st.write(f"**Road:** {row['Road Name']}")
                            st.write(f"**Defect:** {row['Defect Type']}")
                            st.write(f"**Severity:** {row['Severity']}")
                            st.write(f"**Area Affected:** {row['Area (%)']}%")
                            st.write(f"**Recommendation:** {row['Recommendation']}")


# ══════════════════════════════════════════════
# IRI TAB
# ══════════════════════════════════════════════
if mode == "IRI Only":
    iri_tab = tabs[2]
elif mode == "Hybrid (PCI + IRI)":
    iri_tab = tabs[3]
else:
    iri_tab = None

if iri_tab:
    with iri_tab:
        st.markdown("## 📈 IRI Analysis Results")
        if df_iri.empty:
            st.warning("⚠️ No IRI data available. Please upload a file or enter data manually in the Data Input tab.")
        else:
            avg_iri = df_iri["IRI (m/km)"].mean()
            worst_iri = df_iri.loc[df_iri["IRI (m/km)"].idxmax()]
            best_iri = df_iri.loc[df_iri["IRI (m/km)"].idxmin()]
            n = len(df_iri)
            k1, k2, k3, k4 = st.columns(4)
            k1.metric("Average IRI", f"{avg_iri:.2f} m/km")
            k2.metric("Smoothest Section", f"S{int(best_iri['Section'])} ({best_iri['IRI (m/km)']:.2f})")
            k3.metric("Roughest Section", f"S{int(worst_iri['Section'])} ({worst_iri['IRI (m/km)']:.2f})")
            k4.metric("Sections Needing Attention", f"{len(df_iri[df_iri['Condition'].isin(['Poor (Rough)', 'Fair'])])}/{n}")

            st.markdown("---")

            benchmark_table_iri()

            st.markdown("---")
            col1, col2 = st.columns([3, 2])
            with col1:
                st.markdown("### IRI Profile Along Road")
                fig_line = go.Figure()

                fig_line.add_hrect(y0=0, y1=2, fillcolor="#2ecc71", opacity=0.10, line_width=0,
                                   annotation_text="Very Good", annotation_position="left")
                fig_line.add_hrect(y0=2, y1=3, fillcolor="#f1c40f", opacity=0.10, line_width=0,
                                   annotation_text="Good", annotation_position="left")
                fig_line.add_hrect(y0=3, y1=4, fillcolor="#e67e22", opacity=0.10, line_width=0,
                                   annotation_text="Fair", annotation_position="left")
                fig_line.add_hrect(y0=4, y1=max(df_iri["IRI (m/km)"].max() + 1, 6),
                                   fillcolor="#e74c3c", opacity=0.10, line_width=0,
                                   annotation_text="Poor", annotation_position="left")

                fig_line.add_hline(y=2, line_dash="dash", line_color="#2ecc71",
                                   annotation_text="Good threshold (2.0)")
                fig_line.add_hline(y=3, line_dash="dash", line_color="#f39c12",
                                   annotation_text="Fair threshold (3.0)")
                fig_line.add_hline(y=4, line_dash="dash", line_color="#e74c3c",
                                   annotation_text="Poor threshold (4.0)")

                fig_line.add_trace(go.Scatter(
                    x=[f"S{int(r['Section'])}" for _, r in df_iri.iterrows()],
                    y=df_iri["IRI (m/km)"], mode="lines+markers+text",
                    text=[f"{v:.2f}" for v in df_iri["IRI (m/km)"]],
                    textposition="top center",
                    marker=dict(color=[r["Color"] for _, r in df_iri.iterrows()], size=12),
                    line=dict(color="#3498db", width=2)))

                fig_line.update_layout(yaxis_title="IRI (m/km)", xaxis_title="Section",
                                       height=420, plot_bgcolor="rgba(0,0,0,0)", paper_bgcolor="rgba(0,0,0,0)")
                st.plotly_chart(fig_line, use_container_width=True)

            with col2:
                st.markdown("### Condition Distribution")
                iri_cond_counts = df_iri["Condition"].value_counts().reset_index()
                iri_cond_counts.columns = ["Condition", "Count"]
                fig_pie2 = px.pie(iri_cond_counts, names="Condition", values="Count", color="Condition",
                                  color_discrete_map={"Very Good (Smooth)": "#2ecc71", "Good": "#27ae60",
                                                      "Fair": "#f39c12", "Poor (Rough)": "#e74c3c"}, hole=0.4)
                fig_pie2.update_layout(height=420, paper_bgcolor="rgba(0,0,0,0)")
                st.plotly_chart(fig_pie2, use_container_width=True)

            st.markdown("### 📋 IRI Results Table")
            display_iri = df_iri[["Section", "Road Name", "IRI (m/km)", "Condition", "Recommendation"]].copy()
            styled_iri = display_iri.style.map(color_condition, subset=["Condition"])
            st.dataframe(styled_iri, use_container_width=True, hide_index=True)


# ══════════════════════════════════════════════
# DASHBOARD TAB (was "Summary & Hybrid")
# ══════════════════════════════════════════════
if mode == "Hybrid (PCI + IRI)":
    with tabs[4]:
        st.markdown("## 📊 Dashboard")
        if df_pci.empty or df_iri.empty:
            st.warning("⚠️ Both PCI and IRI data are required for the Dashboard.")
        else:
            pci_map = {skey(r["Road Name"], r["Section"]): r for r in pci_results}
            iri_map = {skey(r["Road Name"], r["Section"]): r for r in iri_results}
            common_keys = sorted(set(pci_map.keys()) & set(iri_map.keys()))

            hybrid_rows = []
            for key in common_keys:
                p = pci_map[key]
                r = iri_map[key]
                combined = hybrid_condition(p["Condition"], r["Condition"])
                hybrid_rows.append({
                    "Key": key, "Section": key[1], "Road Name": key[0],
                    "PCI": p["PCI"], "PCI Class": p["Condition"],
                    "IRI (m/km)": r["IRI (m/km)"], "IRI Class": r["Condition"],
                    "Combined Condition": combined, "Maintenance Recommendation": hybrid_recommendation(combined)
                })
            df_hybrid_full = pd.DataFrame(hybrid_rows)

            # ── Section filter ──
            st.markdown("### 🔎 Filter Sections")
            section_labels = {row["Key"]: f"S{row['Section']} — {row['Road Name']}" for row in hybrid_rows}
            selected_keys = st.multiselect(
                "Choose which section(s) to display on the dashboard",
                options=list(section_labels.keys()),
                default=list(section_labels.keys()),
                format_func=lambda x: section_labels[x]
            )

            if not selected_keys:
                st.info("Select at least one section above to view the dashboard.")
            else:
                df_hybrid = df_hybrid_full[df_hybrid_full["Key"].isin(selected_keys)].reset_index(drop=True)

                st.markdown("---")
                k1, k2, k3 = st.columns(3)
                k1.metric("Avg PCI (selected)", f"{df_hybrid['PCI'].mean():.1f}")
                k2.metric("Avg IRI (selected)", f"{df_hybrid['IRI (m/km)'].mean():.2f} m/km")
                critical = len(df_hybrid[df_hybrid["Combined Condition"].isin(["Poor", "Poor (Rough)"])])
                k3.metric("Critical Sections", f"{critical}/{len(df_hybrid)}")

                st.markdown("---")
                st.markdown("### PCI vs IRI Comparison")
                fig_compare = make_subplots(specs=[[{"secondary_y": True}]])
                fig_compare.add_trace(go.Bar(
                    x=[f"S{int(r['Section'])}" for _, r in df_hybrid.iterrows()],
                    y=df_hybrid["PCI"], name="PCI", marker_color="#3498db", opacity=0.8), secondary_y=False)
                fig_compare.add_trace(go.Scatter(
                    x=[f"S{int(r['Section'])}" for _, r in df_hybrid.iterrows()],
                    y=df_hybrid["IRI (m/km)"], name="IRI (m/km)",
                    mode="lines+markers", marker=dict(color="#e74c3c", size=10),
                    line=dict(color="#e74c3c", width=2)), secondary_y=True)

                fig_compare.add_hline(y=85, line_dash="dot", line_color="#2ecc71",
                                      annotation_text="PCI 85 (Very Good)", secondary_y=False)
                fig_compare.add_hline(y=55, line_dash="dot", line_color="#e74c3c",
                                      annotation_text="PCI 55 (Fair)", secondary_y=False)

                fig_compare.update_yaxes(title_text="PCI Score", secondary_y=False)
                fig_compare.update_yaxes(title_text="IRI (m/km)", secondary_y=True)
                fig_compare.update_layout(height=420, plot_bgcolor="rgba(0,0,0,0)", paper_bgcolor="rgba(0,0,0,0)")
                st.plotly_chart(fig_compare, use_container_width=True)

                st.markdown("### 📋 Hybrid Summary Table")
                display_hybrid_cols = ["Section", "Road Name", "PCI", "PCI Class",
                                       "IRI (m/km)", "IRI Class", "Combined Condition", "Maintenance Recommendation"]
                styled_hybrid = df_hybrid[display_hybrid_cols].style.map(
                    color_condition, subset=["PCI Class", "IRI Class", "Combined Condition"])
                st.dataframe(styled_hybrid, use_container_width=True, hide_index=True)

                st.markdown("### 🔧 Maintenance Priority Plan")
                priority_map = {"Poor": 1, "Poor (Rough)": 1, "Fair": 2, "Good / Satisfactory": 3,
                                "Good": 3, "Very Good": 4, "Very Good (Smooth)": 4}
                df_hybrid["Priority"] = df_hybrid["Combined Condition"].map(priority_map)
                df_priority = df_hybrid.sort_values("Priority").reset_index(drop=True)
                df_priority["Priority Level"] = df_priority["Priority"].map(
                    {1: "🔴 Immediate", 2: "🟡 Short-term", 3: "🟢 Scheduled", 4: "✅ Monitor only"})
                for _, row in df_priority.iterrows():
                    key = row["Key"]
                    with st.expander(f"{row['Priority Level']} — Section {key[1]} | {row['Road Name']} | {row['Combined Condition']}"):
                        c1, c2, c3 = st.columns(3)
                        c1.metric("PCI", f"{row['PCI']:.1f}")
                        c2.metric("IRI", f"{row['IRI (m/km)']:.2f} m/km")
                        c3.metric("Combined", row["Combined Condition"])
                        st.info(f"**Recommended Action:** {row['Maintenance Recommendation']}")
                        if key in st.session_state.section_photos:
                            show_section_photo(key, width=320)

                csv = df_hybrid.drop(columns=["Priority", "Key"]).to_csv(index=False)
                st.download_button("⬇️ Download Summary as CSV", data=csv,
                                   file_name="TCG633_Pavement_Results.csv", mime="text/csv")

st.markdown("---")
st.markdown("""
<div style='text-align: center; color: #718096; font-size: 0.85rem;'>
    TCG633 – Digital Pavement Condition Evaluation Tool | UiTM Cawangan Sarawak<br>
    Standards: ASTM D6433 (PCI) | JKR Pavement Maintenance Manual (IRI)
</div>
""", unsafe_allow_html=True)
